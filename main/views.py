from django.shortcuts import render
from director_dashboard.models import Program, Committee, Student
from accounts.models import User

from django.shortcuts import render, redirect
from director_dashboard.models import Program, Committee, Student
from accounts.models import User
from director_dashboard.models import AlbumPhoto

def home(request):
    # Check if user is authenticated
    if request.user.is_authenticated:
        # Redirect based on user role
        if request.user.role == 'director':
            return redirect('dashboard')  # Director dashboard
        elif request.user.role == 'program_manager':
            return redirect('pm_dashboard')  # Program manager dashboard
        elif request.user.role == 'committee_supervisor':
            # REDIRECT BASED ON SUPERVISOR TYPE
            if request.user.supervisor_type == 'cultural':
                return redirect('cultural_dashboard')
            elif request.user.supervisor_type == 'sports':
                return redirect('sports_dashboard')  # You'll create this later
            elif request.user.supervisor_type == 'sharia':
                return redirect('sharia_dashboard')
            elif request.user.supervisor_type == 'scientific':
                return redirect('scientific_dashboard')
            elif request.user.supervisor_type == 'operations':
                return redirect('operations_dashboard')
            else:
                return redirect('home')

    recent_photos = AlbumPhoto.objects.select_related('album').filter(
        album__is_active=True
    ).order_by('-created_at')[:10]


    # If not authenticated, show the public homepage
    context = {
        'total_programs': Program.objects.count(),
        'total_committees': Committee.objects.count(),
        'total_students': Student.objects.count(),
        'total_users': User.objects.count(),
        'recent_photos': recent_photos,
    }
    return render(request, 'main/home.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta
from calendar import monthrange
from accounts.models import User, UserActivity
from director_dashboard.models import Program, Committee
from .models import ScheduleEvent, EventAttendance
from .forms import ScheduleEventForm, EventAttendanceForm, ProgramSelectionForm


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta
from calendar import monthrange

from accounts.models import User, UserActivity
from director_dashboard.models import Program, Committee
from .models import ScheduleEvent, EventAttendance
from .forms import ScheduleEventForm, EventAttendanceForm, ProgramSelectionForm
from pm_dashboard.models import Task, Activity, StudentAttendance
from cultural_committee_dashboard.models import CulturalTask, CulturalReport,TaskSession
from operations_committee_dashboard.models import OperationsTask
from scientific_committee_dashboard.models import ScientificTask, Lecture
from sharia_committee_dashboard.models import ShariaTask, FamilyCompetition
from sports_committee_dashboard.models import SportsTask, Match


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


@login_required
def schedule_calendar(request, program_id=None):
    """Main calendar view for all users with monthly/weekly toggle - includes all committee tasks and recurring tasks"""
    user = request.user

    # Director selects program
    if user.role == 'director':
        if not program_id:
            programs = Program.objects.all()
            if programs.count() == 1:
                return redirect('schedule_calendar', program_id=programs.first().id)

            form = ProgramSelectionForm(request.GET or None)
            if request.GET.get('program'):
                return redirect('schedule_calendar', program_id=request.GET.get('program'))

            context = {
                'form': form,
                'programs': programs,
            }
            return render(request, 'schedule/select_program.html', context)

        program = get_object_or_404(Program, id=program_id)

    # Program Manager
    elif user.role == 'program_manager':
        try:
            program = Program.objects.get(manager=user)
        except Program.DoesNotExist:
            messages.error(request, 'لم يتم تعيين برنامج لك بعد')
            return redirect('home')

    # Committee Supervisor
    elif user.role == 'committee_supervisor':
        try:
            committee = Committee.objects.get(supervisor=user)
            program = committee.program
        except Committee.DoesNotExist:
            messages.error(request, 'لم يتم تعيين لجنة لك بعد')
            return redirect('home')

    else:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('home')

    # Get view type (monthly or weekly)
    view_type = request.GET.get('view', 'monthly')

    # Get month, year, and week from query params
    # Get month, year, and week from query params
    now = timezone.now()

    # Handle year parameter
    year_param = request.GET.get('year')
    if year_param and year_param != 'None':
        year = int(year_param)
    else:
        year = now.year

    # Handle month parameter
    month_param = request.GET.get('month')
    if month_param and month_param != 'None':
        month = int(month_param)
    else:
        month = now.month

    # Handle week parameter
    week_param = request.GET.get('week')
    if week_param and week_param != 'None':
        week_number = int(week_param)
    else:
        week_number = now.isocalendar()[1]

    # Calculate previous and next month
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year

    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    if view_type == 'weekly':
        # Weekly view logic
        monday_of_week = datetime.strptime(f'{year}-W{week_number:02d}-1', "%Y-W%W-%w").date()

        # Adjust to Saturday (go back 2 days from Monday)
        week_start = monday_of_week - timedelta(days=2)
        week_end = week_start + timedelta(days=6)

        prev_week_start = week_start - timedelta(days=7)
        prev_week = prev_week_start.isocalendar()[1]
        prev_week_year = prev_week_start.year

        next_week_start = week_start + timedelta(days=7)
        next_week = next_week_start.isocalendar()[1]
        next_week_year = next_week_start.year

        start_date = week_start
        end_date = week_end

    else:
        # Monthly view logic
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, month, monthrange(year, month)[1]).date()

    event_type_filter = request.GET.get('event_type', '')  # Now this is the unified filter
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    committee_filter = request.GET.get('committee', '')

    # Get all events and tasks for the period
    events = ScheduleEvent.objects.filter(
        program=program,
        start_date__gte=start_date,
        start_date__lte=end_date
    ).select_related('committee', 'created_by').order_by('start_date', 'start_time')

    if status_filter:
        events = events.filter(status=status_filter)
    if priority_filter:
        events = events.filter(priority=priority_filter)
    if committee_filter:
        events = events.filter(committee_id=committee_filter)

    # Get regular tasks - now handling recurring tasks
    tasks = Task.objects.filter(
        program=program
    ).select_related('committee')

    if committee_filter:
        tasks = tasks.filter(committee_id=committee_filter)

    # Process recurring tasks
    task_occurrences = {}  # {date: [(task, is_start, is_end, group_id), ...]}

    for task in tasks:
        if task.is_recurring:
            # IMPORTANT: Only get occurrences within the view's date range
            # Don't process if task hasn't started yet
            task_start = task.start_date or task.due_date
            if task_start > end_date:
                continue

            # Don't process if task has ended
            if task.recurrence_end_date and task.recurrence_end_date < start_date:
                continue

            # Get consecutive day groups for this task within the view range
            groups = task.get_consecutive_day_groups(start_date, end_date)
            for group_idx, (group_start, group_end) in enumerate(groups):
                group_id = f"task_{task.id}_group_{group_idx}"

                # Add all dates in this group
                current = group_start
                while current <= group_end:
                    if start_date <= current <= end_date:
                        if current not in task_occurrences:
                            task_occurrences[current] = []

                        task_occurrences[current].append({
                            'task': task,
                            'is_start': current == group_start,
                            'is_end': current == group_end,
                            'group_id': group_id,
                            'group_start': group_start,
                            'group_end': group_end,
                            'span_days': (group_end - group_start).days + 1
                        })
                    current += timedelta(days=1)
        else:
            # Non-recurring task
            if start_date <= task.due_date <= end_date:
                if task.due_date not in task_occurrences:
                    task_occurrences[task.due_date] = []
                task_occurrences[task.due_date].append({
                    'task': task,
                    'is_start': True,
                    'is_end': True,
                    'group_id': f"task_{task.id}_single",
                    'group_start': task.due_date,
                    'group_end': task.due_date,
                    'span_days': 1
                })

    cultural_tasks = CulturalTask.objects.filter(
        committee__program=program
    ).select_related('committee')

    if committee_filter:
        cultural_tasks = cultural_tasks.filter(committee_id=committee_filter)

    # Process recurring cultural tasks
    cultural_task_occurrences = {}  # {date: [(cultural_task, is_start, is_end, group_id), ...]}

    for cultural_task in cultural_tasks:
        if cultural_task.is_recurring:
            # IMPORTANT: Only get occurrences within the view's date range
            # Don't process if task hasn't started yet
            cultural_task_start = cultural_task.start_date or cultural_task.due_date
            if cultural_task_start > end_date:
                continue

            # Don't process if task has ended
            if cultural_task.recurrence_end_date and cultural_task.recurrence_end_date < start_date:
                continue

            # Get consecutive day groups for this task within the view range
            groups = cultural_task.get_consecutive_day_groups(start_date, end_date)
            for group_idx, (group_start, group_end) in enumerate(groups):
                group_id = f"cultural_task_{cultural_task.id}_group_{group_idx}"

                # Add all dates in this group
                current = group_start
                while current <= group_end:
                    if start_date <= current <= end_date:
                        if current not in cultural_task_occurrences:
                            cultural_task_occurrences[current] = []

                        cultural_task_occurrences[current].append({
                            'task': cultural_task,
                            'is_start': current == group_start,
                            'is_end': current == group_end,
                            'group_id': group_id,
                            'group_start': group_start,
                            'group_end': group_end,
                            'span_days': (group_end - group_start).days + 1,
                            'type': 'cultural_task'  # Add type identifier
                        })
                    current += timedelta(days=1)
        else:
            # Non-recurring cultural task
            if start_date <= cultural_task.due_date <= end_date:
                if cultural_task.due_date not in cultural_task_occurrences:
                    cultural_task_occurrences[cultural_task.due_date] = []
                cultural_task_occurrences[cultural_task.due_date].append({
                    'task': cultural_task,
                    'is_start': True,
                    'is_end': True,
                    'group_id': f"cultural_task_{cultural_task.id}_single",
                    'group_start': cultural_task.due_date,
                    'group_end': cultural_task.due_date,
                    'span_days': 1,
                    'type': 'cultural_task'  # Add type identifier
                })

    # Combine all task occurrences
    all_task_occurrences = {}
    for date, occurrences in task_occurrences.items():
        if date not in all_task_occurrences:
            all_task_occurrences[date] = []
        all_task_occurrences[date].extend(occurrences)

    for date, occurrences in cultural_task_occurrences.items():
        if date not in all_task_occurrences:
            all_task_occurrences[date] = []
        all_task_occurrences[date].extend(occurrences)

    activities = Activity.objects.filter(
        program=program,
        date__lte=end_date,
        date__gte=start_date
    ).select_related('committee', 'created_by')

    cultural_tasks_non_recurring = CulturalTask.objects.filter(
        committee__program=program,
        is_recurring=False,
        due_date__lte=end_date,
        due_date__gte=start_date
    ).select_related('committee')

    task_sessions = TaskSession.objects.filter(
        task__committee__program=program,
        date__lte=end_date,
        date__gte=start_date
    ).select_related('task', 'task__committee')

    operations_tasks = OperationsTask.objects.filter(
        committee__program=program,
        due_date__lte=end_date,
        due_date__gte=start_date
    ).select_related('committee')

    scientific_tasks_all = ScientificTask.objects.filter(
        committee__program=program
    ).select_related('committee')

    if committee_filter:
        scientific_tasks_all = scientific_tasks_all.filter(committee_id=committee_filter)

    # Process recurring scientific tasks
    scientific_task_occurrences = {}  # {date: [(scientific_task, is_start, is_end, group_id), ...]}

    for scientific_task in scientific_tasks_all:
        if scientific_task.is_recurring:
            # IMPORTANT: Only get occurrences within the view's date range
            # Don't process if task hasn't started yet
            scientific_task_start = scientific_task.start_date or scientific_task.due_date
            if scientific_task_start > end_date:
                continue

            # Don't process if task has ended
            if scientific_task.recurrence_end_date and scientific_task.recurrence_end_date < start_date:
                continue

            # Get consecutive day groups for this task within the view range
            groups = scientific_task.get_consecutive_day_groups(start_date, end_date)
            for group_idx, (group_start, group_end) in enumerate(groups):
                group_id = f"scientific_task_{scientific_task.id}_group_{group_idx}"

                # Add all dates in this group
                current = group_start
                while current <= group_end:
                    if start_date <= current <= end_date:
                        if current not in scientific_task_occurrences:
                            scientific_task_occurrences[current] = []

                        scientific_task_occurrences[current].append({
                            'task': scientific_task,
                            'is_start': current == group_start,
                            'is_end': current == group_end,
                            'group_id': group_id,
                            'group_start': group_start,
                            'group_end': group_end,
                            'span_days': (group_end - group_start).days + 1,
                            'type': 'scientific_task'  # Add type identifier
                        })
                    current += timedelta(days=1)
        else:
            # Non-recurring scientific task
            if start_date <= scientific_task.due_date <= end_date:
                if scientific_task.due_date not in scientific_task_occurrences:
                    scientific_task_occurrences[scientific_task.due_date] = []
                scientific_task_occurrences[scientific_task.due_date].append({
                    'task': scientific_task,
                    'is_start': True,
                    'is_end': True,
                    'group_id': f"scientific_task_{scientific_task.id}_single",
                    'group_start': scientific_task.due_date,
                    'group_end': scientific_task.due_date,
                    'span_days': 1,
                    'type': 'scientific_task'  # Add type identifier
                })

    # Add scientific task occurrences to all_task_occurrences
    for date, occurrences in scientific_task_occurrences.items():
        if date not in all_task_occurrences:
            all_task_occurrences[date] = []
        all_task_occurrences[date].extend(occurrences)

    # Update the scientific_tasks queryset to only get non-recurring tasks for display
    scientific_tasks = ScientificTask.objects.filter(
        committee__program=program,
        is_recurring=False,  # Only non-recurring tasks
        due_date__lte=end_date,
        due_date__gte=start_date
    ).select_related('committee')

    lectures = Lecture.objects.filter(
        committee__program=program,
        date__lte=end_date,
        date__gte=start_date
    ).select_related('committee', 'created_by')

    sharia_tasks_all = ShariaTask.objects.filter(
        committee__program=program
    ).select_related('committee')

    if committee_filter:
        sharia_tasks_all = sharia_tasks_all.filter(committee_id=committee_filter)

    # Process recurring sharia tasks
    sharia_task_occurrences = {}  # {date: [(sharia_task, is_start, is_end, group_id), ...]}

    for sharia_task in sharia_tasks_all:
        if sharia_task.is_recurring:
            # IMPORTANT: Only get occurrences within the view's date range
            # Don't process if task hasn't started yet
            sharia_task_start = sharia_task.start_date or sharia_task.due_date
            if sharia_task_start > end_date:
                continue

            # Don't process if task has ended
            if sharia_task.recurrence_end_date and sharia_task.recurrence_end_date < start_date:
                continue

            # Get consecutive day groups for this task within the view range
            groups = sharia_task.get_consecutive_day_groups(start_date, end_date)
            for group_idx, (group_start, group_end) in enumerate(groups):
                group_id = f"sharia_task_{sharia_task.id}_group_{group_idx}"

                # Add all dates in this group
                current = group_start
                while current <= group_end:
                    if start_date <= current <= end_date:
                        if current not in sharia_task_occurrences:
                            sharia_task_occurrences[current] = []

                        sharia_task_occurrences[current].append({
                            'task': sharia_task,
                            'is_start': current == group_start,
                            'is_end': current == group_end,
                            'group_id': group_id,
                            'group_start': group_start,
                            'group_end': group_end,
                            'span_days': (group_end - group_start).days + 1,
                            'type': 'sharia_task'  # Add type identifier
                        })
                    current += timedelta(days=1)
        else:
            # Non-recurring sharia task
            if start_date <= sharia_task.due_date <= end_date:
                if sharia_task.due_date not in sharia_task_occurrences:
                    sharia_task_occurrences[sharia_task.due_date] = []
                sharia_task_occurrences[sharia_task.due_date].append({
                    'task': sharia_task,
                    'is_start': True,
                    'is_end': True,
                    'group_id': f"sharia_task_{sharia_task.id}_single",
                    'group_start': sharia_task.due_date,
                    'group_end': sharia_task.due_date,
                    'span_days': 1,
                    'type': 'sharia_task'  # Add type identifier
                })

    for date, occurrences in sharia_task_occurrences.items():
        if date not in all_task_occurrences:
            all_task_occurrences[date] = []
        all_task_occurrences[date].extend(occurrences)

    sharia_tasks_non_recurring = ShariaTask.objects.filter(
        committee__program=program,
        is_recurring=False,
        due_date__lte=end_date,
        due_date__gte=start_date
    ).select_related('committee')

    family_competitions = FamilyCompetition.objects.filter(
        committee__program=program,
        start_date__lte=end_date,
        end_date__gte=start_date
    ).select_related('committee')

    sports_tasks = SportsTask.objects.filter(
        committee__program=program
    ).select_related('committee')

    if committee_filter:
        sports_tasks = sports_tasks.filter(committee_id=committee_filter)

    # Process recurring sports tasks
    sports_task_occurrences = {}  # {date: [(sports_task, is_start, is_end, group_id), ...]}

    for sports_task in sports_tasks:
        if sports_task.is_recurring:
            # IMPORTANT: Only get occurrences within the view's date range
            # Don't process if task hasn't started yet
            sports_task_start = sports_task.start_date or sports_task.due_date
            if sports_task_start > end_date:
                continue

            # Don't process if task has ended
            if sports_task.recurrence_end_date and sports_task.recurrence_end_date < start_date:
                continue

            # Get consecutive day groups for this task within the view range
            groups = sports_task.get_consecutive_day_groups(start_date, end_date)
            for group_idx, (group_start, group_end) in enumerate(groups):
                group_id = f"sports_task_{sports_task.id}_group_{group_idx}"

                # Add all dates in this group
                current = group_start
                while current <= group_end:
                    if start_date <= current <= end_date:
                        if current not in sports_task_occurrences:
                            sports_task_occurrences[current] = []

                        sports_task_occurrences[current].append({
                            'task': sports_task,
                            'is_start': current == group_start,
                            'is_end': current == group_end,
                            'group_id': group_id,
                            'group_start': group_start,
                            'group_end': group_end,
                            'span_days': (group_end - group_start).days + 1,
                            'type': 'sports_task'  # Add type identifier
                        })
                    current += timedelta(days=1)
        else:
            # Non-recurring sports task
            if start_date <= sports_task.due_date <= end_date:
                if sports_task.due_date not in sports_task_occurrences:
                    sports_task_occurrences[sports_task.due_date] = []
                sports_task_occurrences[sports_task.due_date].append({
                    'task': sports_task,
                    'is_start': True,
                    'is_end': True,
                    'group_id': f"sports_task_{sports_task.id}_single",
                    'group_start': sports_task.due_date,
                    'group_end': sports_task.due_date,
                    'span_days': 1,
                    'type': 'sports_task'  # Add type identifier
                })

    for date, occurrences in sports_task_occurrences.items():
        if date not in all_task_occurrences:
            all_task_occurrences[date] = []
        all_task_occurrences[date].extend(occurrences)

    matches = Match.objects.filter(
        committee__program=program,
        date__lte=end_date,
        date__gte=start_date
    ).select_related('committee', 'created_by')

    # ========== APPLY STATUS FILTER TO ALL TYPES THAT HAVE STATUS FIELD ==========
    if status_filter:
        # Models with status field: ScheduleEvent, Task, CulturalTask, OperationsTask,
        # ScientificTask, ShariaTask, SportsTask, Lecture, Match, FamilyCompetition
        events = events.filter(status=status_filter)
        tasks = tasks.filter(status=status_filter)
        cultural_tasks = cultural_tasks.filter(status=status_filter)
        operations_tasks = operations_tasks.filter(status=status_filter)
        scientific_tasks_all = scientific_tasks_all.filter(status=status_filter)
        sharia_tasks_all = sharia_tasks_all.filter(status=status_filter)
        sports_tasks = sports_tasks.filter(status=status_filter)
        lectures = lectures.filter(status=status_filter)
        matches = matches.filter(status=status_filter)
        family_competitions = family_competitions.filter(status=status_filter)

        for date in list(all_task_occurrences.keys()):
            filtered_occurrences = []
            for occ in all_task_occurrences[date]:
                if occ['type'] == 'regular_task' and occ['task'].status == status_filter:
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'cultural_task' and occ['task'].status == status_filter:
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'scientific_task' and occ['task'].status == status_filter:
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'sharia_task' and occ['task'].status == status_filter:
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'sports_task' and occ['task'].status == status_filter:
                    filtered_occurrences.append(occ)
            if filtered_occurrences:
                all_task_occurrences[date] = filtered_occurrences
            else:
                del all_task_occurrences[date]

    # ========== APPLY PRIORITY FILTER TO ALL TYPES THAT HAVE PRIORITY FIELD ==========
    if priority_filter:
        # Models with priority field: ScheduleEvent, Task, OperationsTask
        events = events.filter(priority=priority_filter)
        tasks = tasks.filter(priority=priority_filter)
        cultural_tasks_non_recurring = cultural_tasks_non_recurring.filter(priority=priority_filter)
        operations_tasks = operations_tasks.filter(priority=priority_filter)
        scientific_tasks_all = scientific_tasks_all.filter(priority=priority_filter)

        for date in list(all_task_occurrences.keys()):
            filtered_occurrences = []
            for occ in all_task_occurrences[date]:
                if occ['type'] == 'regular_task' and occ['task'].priority == priority_filter:
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'cultural_task' and occ['task'].priority == priority_filter:
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'sharia_task' and occ['task'].priority == priority_filter:
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'scientific_task' and occ['task'].priority == priority_filter:
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'sports_task' and occ['task'].priority == priority_filter:
                    filtered_occurrences.append(occ)
            if filtered_occurrences:
                all_task_occurrences[date] = filtered_occurrences
            else:
                del all_task_occurrences[date]

    # ========== APPLY COMMITTEE FILTER TO ALL EVENT TYPES ==========
    if committee_filter:
        events = events.filter(committee_id=committee_filter)
        tasks = tasks.filter(committee_id=committee_filter)
        activities = activities.filter(committee_id=committee_filter)
        cultural_tasks = cultural_tasks.filter(committee_id=committee_filter)
        task_sessions = task_sessions.filter(task__committee_id=committee_filter)
        operations_tasks = operations_tasks.filter(committee_id=committee_filter)
        scientific_tasks_all = scientific_tasks_all.filter(committee_id=committee_filter)
        lectures = lectures.filter(committee_id=committee_filter)
        sharia_tasks_all = sharia_tasks_all.filter(committee_id=committee_filter)
        family_competitions = family_competitions.filter(committee_id=committee_filter)
        sports_tasks = sports_tasks.filter(committee_id=committee_filter)
        matches = matches.filter(committee_id=committee_filter)

        for date in list(all_task_occurrences.keys()):
            filtered_occurrences = []
            for occ in all_task_occurrences[date]:
                if occ['type'] == 'regular_task' and occ['task'].committee_id == int(committee_filter):
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'cultural_task' and occ['task'].committee_id == int(committee_filter):
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'sharia_task' and occ['task'].committee_id == int(committee_filter):
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'scientific_task' and occ['task'].committee_id == int(committee_filter):
                    filtered_occurrences.append(occ)
                elif occ['type'] == 'sports_task' and occ['task'].committee_id == int(committee_filter):
                    filtered_occurrences.append(occ)
            if filtered_occurrences:
                all_task_occurrences[date] = filtered_occurrences
            else:
                del all_task_occurrences[date]

    if event_type_filter:
        if event_type_filter == 'schedule_event':
            # Keep only ScheduleEvents, clear all others
            activities = activities.none()
            cultural_tasks = cultural_tasks.none()
            task_sessions = task_sessions.none()
            operations_tasks = operations_tasks.none()
            scientific_tasks = scientific_tasks.none()
            lectures = lectures.none()
            sharia_tasks_all = sharia_tasks_all.none()
            family_competitions = family_competitions.none()
            sports_tasks = sports_tasks.none()
            matches = matches.none()
            task_occurrences = {}
        elif event_type_filter == 'task':
            # Keep only Tasks
            events = events.none()
            activities = activities.none()
            cultural_tasks = cultural_tasks.none()
            task_sessions = task_sessions.none()
            operations_tasks = operations_tasks.none()
            scientific_tasks = scientific_tasks.none()
            lectures = lectures.none()
            sharia_tasks_all = sharia_tasks_all.none()
            family_competitions = family_competitions.none()
            sports_tasks = sports_tasks.none()
            matches = matches.none()
        elif event_type_filter == 'activity':
            # Keep only Activities
            events = events.none()
            cultural_tasks = cultural_tasks.none()
            task_sessions = task_sessions.none()
            operations_tasks = operations_tasks.none()
            scientific_tasks = scientific_tasks.none()
            lectures = lectures.none()
            sharia_tasks_all = sharia_tasks_all.none()
            family_competitions = family_competitions.none()
            sports_tasks = sports_tasks.none()
            matches = matches.none()
            task_occurrences = {}
        elif event_type_filter == 'cultural_task':
            # Keep only Cultural Tasks
            events = events.none()
            activities = activities.none()
            task_sessions = task_sessions.none()
            operations_tasks = operations_tasks.none()
            scientific_tasks = scientific_tasks.none()
            lectures = lectures.none()
            sharia_tasks_all = sharia_tasks_all.none()
            family_competitions = family_competitions.none()
            sports_tasks = sports_tasks.none()
            matches = matches.none()
            task_occurrences = {}
        elif event_type_filter == 'task_session':
            # Keep only Task Sessions
            events = events.none()
            activities = activities.none()
            cultural_tasks = cultural_tasks.none()
            operations_tasks = operations_tasks.none()
            scientific_tasks = scientific_tasks.none()
            lectures = lectures.none()
            sharia_tasks_all = sharia_tasks_all.none()
            family_competitions = family_competitions.none()
            sports_tasks = sports_tasks.none()
            matches = matches.none()
            task_occurrences = {}
        elif event_type_filter == 'operations_task':
            # Keep only Operations Tasks
            events = events.none()
            activities = activities.none()
            cultural_tasks = cultural_tasks.none()
            task_sessions = task_sessions.none()
            scientific_tasks = scientific_tasks.none()
            lectures = lectures.none()
            sharia_tasks_all = sharia_tasks_all.none()
            family_competitions = family_competitions.none()
            sports_tasks = sports_tasks.none()
            matches = matches.none()
            task_occurrences = {}
        elif event_type_filter == 'scientific_task':
            # Keep only Scientific Tasks
            events = events.none()
            activities = activities.none()
            cultural_tasks = cultural_tasks.none()
            task_sessions = task_sessions.none()
            operations_tasks = operations_tasks.none()
            lectures = lectures.none()
            sharia_tasks_all = sharia_tasks_all.none()
            family_competitions = family_competitions.none()
            sports_tasks = sports_tasks.none()
            matches = matches.none()

            # Filter all_task_occurrences to only include scientific tasks
            for date in list(all_task_occurrences.keys()):
                filtered_occurrences = []
                for occ in all_task_occurrences[date]:
                    if occ['type'] == 'scientific_task':
                        filtered_occurrences.append(occ)
                if filtered_occurrences:
                    all_task_occurrences[date] = filtered_occurrences
                else:
                    del all_task_occurrences[date]
        elif event_type_filter == 'lecture':
            # Keep only Lectures
            events = events.none()
            activities = activities.none()
            cultural_tasks = cultural_tasks.none()
            task_sessions = task_sessions.none()
            operations_tasks = operations_tasks.none()
            scientific_tasks = scientific_tasks.none()
            sharia_tasks_all = sharia_tasks_all.none()
            family_competitions = family_competitions.none()
            sports_tasks = sports_tasks.none()
            matches = matches.none()
            task_occurrences = {}
        elif event_type_filter == 'sharia_task':
            # Keep only Sharia Tasks
            events = events.none()
            activities = activities.none()
            cultural_tasks = cultural_tasks.none()
            task_sessions = task_sessions.none()
            operations_tasks = operations_tasks.none()
            scientific_tasks = scientific_tasks.none()
            lectures = lectures.none()
            family_competitions = family_competitions.none()
            sports_tasks = sports_tasks.none()
            matches = matches.none()
            task_occurrences = {}
        elif event_type_filter == 'family_competition':
            # Keep only Family Competitions
            events = events.none()
            activities = activities.none()
            cultural_tasks = cultural_tasks.none()
            task_sessions = task_sessions.none()
            operations_tasks = operations_tasks.none()
            scientific_tasks = scientific_tasks.none()
            lectures = lectures.none()
            sharia_tasks_all = sharia_tasks_all.none()
            sports_tasks = sports_tasks.none()
            matches = matches.none()
            task_occurrences = {}
        elif event_type_filter == 'sports_task':
            # Keep only Sports Tasks
            events = events.none()
            activities = activities.none()
            cultural_tasks = cultural_tasks.none()
            task_sessions = task_sessions.none()
            operations_tasks = operations_tasks.none()
            scientific_tasks = scientific_tasks.none()
            lectures = lectures.none()
            sharia_tasks_all = sharia_tasks_all.none()
            family_competitions = family_competitions.none()
            matches = matches.none()
            task_occurrences = {}
        elif event_type_filter == 'match':
            # Keep only Matches
            events = events.none()
            activities = activities.none()
            cultural_tasks = cultural_tasks.none()
            task_sessions = task_sessions.none()
            operations_tasks = operations_tasks.none()
            scientific_tasks = scientific_tasks.none()
            lectures = lectures.none()
            sharia_tasks_all = sharia_tasks_all.none()
            family_competitions = family_competitions.none()
            sports_tasks = sports_tasks.none()
            task_occurrences = {}

    # Filter by committee if supervisor
    if user.role == 'committee_supervisor':
        events = events.filter(Q(committee=committee) | Q(committee__isnull=True))
        activities = activities.filter(Q(committee=committee) | Q(committee__isnull=True))
        cultural_tasks = [ct for ct in cultural_tasks if ct.committee == committee]
        task_sessions = task_sessions.filter(task__committee=committee)
        operations_tasks = operations_tasks.filter(committee=committee)
        scientific_tasks = scientific_tasks.filter(committee=committee)
        lectures = lectures.filter(committee=committee)
        sharia_tasks_all = [st for st in sharia_tasks_all if st.committee == committee]
        family_competitions = family_competitions.filter(committee=committee)
        sports_tasks = sports_tasks.filter(committee=committee)
        matches = matches.filter(committee=committee)

    if view_type == 'weekly':
        # Build week days
        week_days = []
        for i in range(7):
            day_date = week_start + timedelta(days=i)
            day_events = events.filter(start_date=day_date)
            day_task_occurrences = all_task_occurrences.get(day_date, [])
            day_activities = activities.filter(date=day_date)
            day_cultural_tasks = cultural_tasks_non_recurring.filter(due_date=day_date)
            day_task_sessions = task_sessions.filter(date=day_date)
            day_operations_tasks = operations_tasks.filter(due_date=day_date)
            day_scientific_tasks = scientific_tasks.filter(due_date=day_date)
            day_lectures = lectures.filter(date=day_date)
            day_sharia_tasks = sharia_tasks_non_recurring.filter(due_date=day_date)
            day_family_competitions = family_competitions.filter(
                start_date__lte=day_date,
                end_date__gte=day_date
            )
            day_sports_tasks = sports_tasks.filter(due_date=day_date)
            day_matches = matches.filter(date=day_date)

            total_count = (
                    day_events.count() + len(day_task_occurrences) + day_activities.count() +
                    day_cultural_tasks.count() + day_task_sessions.count() + day_operations_tasks.count() +
                    day_scientific_tasks.count() + day_lectures.count() +
                    day_sharia_tasks.count() + day_family_competitions.count() +
                    day_sports_tasks.count() + day_matches.count()
            )

            week_days.append({
                'date': day_date,
                'day': day_date.day,
                'is_today': day_date == now.date(),
                'is_current_month': True,
                'events': day_events,
                'task_occurrences': day_task_occurrences,
                'activities': day_activities,
                'cultural_tasks': day_cultural_tasks,
                'operations_tasks': day_operations_tasks,
                'task_sessions': day_task_sessions,
                'scientific_tasks': day_scientific_tasks,
                'lectures': day_lectures,
                'sharia_tasks': day_sharia_tasks,
                'family_competitions': day_family_competitions,
                'sports_tasks': day_sports_tasks,
                'matches': day_matches,
                'total_events': total_count,
            })

        weeks = [week_days]
        calendar_data = [week_days]

        month_names = [
            'يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',
            'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'
        ]

        context_extra = {
            'view_type': 'weekly',
            'week_number': week_number,
            'week_start': week_start,
            'week_end': week_end,
            'prev_week': prev_week,
            'prev_week_year': prev_week_year,
            'next_week': next_week,
            'next_week_year': next_week_year,
            'month_name': f"{month_names[week_start.month - 1]} - {month_names[week_end.month - 1]}" if week_start.month != week_end.month else
            month_names[week_start.month - 1],
        }



    else:

        # Monthly view logic

        first_day = datetime(year, month, 1).date()

        last_day = datetime(year, month, monthrange(year, month)[1]).date()

        cal_data = []

        # Get weekday of first day (0=Monday, 6=Sunday in Python)

        # But we want Saturday as first column (0=Saturday, 6=Friday)

        first_weekday = first_day.weekday()  # Monday=0, Sunday=6

        # Convert to our calendar where Saturday is first column (0)

        # Python: Mon(0), Tue(1), Wed(2), Thu(3), Fri(4), Sat(5), Sun(6)

        # We want: Sat(0), Sun(1), Mon(2), Tue(3), Wed(4), Thu(5), Fri(6)

        if first_weekday == 5:  # Saturday

            first_weekday = 0

        elif first_weekday == 6:  # Sunday

            first_weekday = 1

        elif first_weekday == 0:  # Monday

            first_weekday = 2

        elif first_weekday == 1:  # Tuesday

            first_weekday = 3

        elif first_weekday == 2:  # Wednesday

            first_weekday = 4

        elif first_weekday == 3:  # Thursday

            first_weekday = 5

        elif first_weekday == 4:  # Friday

            first_weekday = 6

        # Add blank days from previous month

        if first_weekday > 0:

            prev_month_days = monthrange(prev_year, prev_month)[1]

            for i in range(first_weekday):
                day = prev_month_days - first_weekday + i + 1

                date = datetime(prev_year, prev_month, day).date()

                cal_data.append({

                    'day': day,

                    'date': date,

                    'is_current_month': False,

                    'is_today': date == now.date(),

                    'events': [],

                    'task_occurrences': [],

                    'activities': [],

                    'cultural_tasks': [],

                    'operations_tasks': [],

                    'scientific_tasks': [],

                    'task_sessions': [],

                    'lectures': [],

                    'sharia_tasks': [],

                    'family_competitions': [],

                    'sports_tasks': [],

                    'matches': [],

                    'total_events': 0,

                })

        days_in_month = monthrange(year, month)[1]

        for day in range(1, days_in_month + 1):

            date = datetime(year, month, day).date()

            # Determine if weekend (Friday = 4, Saturday = 5 in Python weekday)

            # In our adjusted calendar: Saturday=0, Sunday=1, Friday=6

            python_weekday = date.weekday()  # Monday=0, Sunday=6

            is_weekend = False

            if python_weekday == 4:  # Friday

                is_weekend = True

            elif python_weekday == 5:  # Saturday

                is_weekend = True

            day_events = events.filter(start_date=date)

            day_task_occurrences = all_task_occurrences.get(date, [])

            day_activities = activities.filter(date=date)

            day_cultural_tasks = cultural_tasks_non_recurring.filter(due_date=date)

            day_task_sessions = task_sessions.filter(date=date)

            day_operations_tasks = operations_tasks.filter(due_date=date)

            day_scientific_tasks = scientific_tasks.filter(due_date=date)

            day_lectures = lectures.filter(date=date)

            day_sharia_tasks = sharia_tasks_non_recurring.filter(due_date=date)

            day_family_competitions = family_competitions.filter(

                start_date__lte=date,

                end_date__gte=date

            )

            day_sports_tasks = SportsTask.objects.filter(
                committee__program=program,
                is_recurring=False,  # Only non-recurring tasks
                due_date=date
            )
            if committee_filter:
                day_sports_tasks = day_sports_tasks.filter(committee_id=committee_filter)

            day_matches = matches.filter(date=date)

            total_count = (

                    day_events.count() + len(day_task_occurrences) + day_activities.count() +

                    day_cultural_tasks.count() + day_task_sessions.count() + day_operations_tasks.count() +

                    day_scientific_tasks.count() + day_lectures.count() +

                    day_sharia_tasks.count() + day_family_competitions.count() +

                    day_sports_tasks.count() + day_matches.count()

            )

            cal_data.append({

                'day': day,

                'date': date,

                'is_current_month': True,

                'is_today': date == now.date(),

                'is_weekend': is_weekend,

                'events': day_events,

                'task_occurrences': day_task_occurrences,

                'activities': day_activities,

                'cultural_tasks': day_cultural_tasks,

                'operations_tasks': day_operations_tasks,

                'scientific_tasks': day_scientific_tasks,

                'task_sessions': day_task_sessions,

                'lectures': day_lectures,

                'sharia_tasks': day_sharia_tasks,

                'family_competitions': day_family_competitions,

                'sports_tasks': day_sports_tasks,

                'matches': day_matches,

                'total_events': total_count,

            })

        # Fill remaining cells

        total_cells_needed = 42  # 6 weeks * 7 days

        remaining = total_cells_needed - len(cal_data)

        for day in range(1, remaining + 1):
            date = datetime(next_year, next_month, day).date()

            cal_data.append({

                'day': day,

                'date': date,

                'is_current_month': False,

                'is_today': date == now.date(),

                'events': [],

                'task_occurrences': [],

                'activities': [],

                'cultural_tasks': [],

                'operations_tasks': [],

                'scientific_tasks': [],

                'task_sessions': [],

                'lectures': [],

                'sharia_tasks': [],

                'family_competitions': [],

                'sports_tasks': [],

                'matches': [],

                'total_events': 0,

            })

        # Split into weeks (7 days each)

        weeks = []

        for i in range(0, len(cal_data), 7):
            weeks.append(cal_data[i:i + 7])

        calendar_data = weeks

        month_names = [

            'يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو',

            'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر'

        ]

        month_name = month_names[month - 1]

        context_extra = {

            'view_type': 'monthly',

            'month_name': month_name,

            'prev_year': prev_year,

            'prev_month': prev_month,

            'next_year': next_year,

            'next_month': next_month,

        }

    # Determine base template
    if user.role == 'director':
        base_template = 'director_base.html'
    elif user.role == 'program_manager':
        base_template = 'program_manager_base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'cultural':
        base_template = 'cultural_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'sports':
        base_template = 'sports_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'sharia':
        base_template = 'sharia_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'operations':
        base_template = 'operations_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'scientific':
        base_template = 'scientific_committee/base.html'
    else:
        base_template = 'base.html'

    context = {
        'program': program,
        'weeks': weeks,
        'calendar_data': calendar_data,
        'view_type': view_type,
        'month_name': month_names[month-1] if 1 <= month <= 12 else '',
        'week_number': week_number if view_type == 'weekly' else None,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
        'prev_week': prev_week if view_type == 'weekly' else None,
        'prev_week_year': prev_week_year if view_type == 'weekly' else None,
        'next_week': next_week if view_type == 'weekly' else None,
        'next_week_year': next_week_year if view_type == 'weekly' else None,
        'year': year,
        'month': month,
        'today': now.date(),
        'base_template': base_template,
    }
    context.update(context_extra)

    return render(request, 'schedule/calendar.html', context)


@login_required
def day_events(request, program_id, year, month, day):
    """Show all events for a specific day - includes all committee tasks and recurring tasks"""
    program = get_object_or_404(Program, id=program_id)
    user = request.user

    # Check permissions
    if not has_permission_for_program(user, program):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('home')

    date = datetime(year, month, day).date()

    # Get all events for this day
    events = ScheduleEvent.objects.filter(
        program=program,
        start_date=date
    ).select_related('committee', 'created_by').order_by('start_time')

    # Get ALL tasks (including recurring ones)
    all_tasks = Task.objects.filter(program=program).select_related('committee')

    all_cultural_tasks = CulturalTask.objects.filter(
        committee__program=program
    ).select_related('committee')

    all_sports_tasks = SportsTask.objects.filter(
        committee__program=program
    ).select_related('committee')

    all_sharia_tasks = ShariaTask.objects.filter(
        committee__program=program
    ).select_related('committee')

    all_scientific_tasks = ScientificTask.objects.filter(
        committee__program=program
    ).select_related('committee')

    # Filter tasks that occur on this date
    tasks_for_day = []
    for task in all_tasks:
        if task.is_recurring:
            # Check if this date is in the task's occurrence dates
            occurrences = task.get_occurrence_dates(date, date)
            if date in occurrences:
                tasks_for_day.append({
                    'task': task,
                    'type': 'regular_task'
                })
        else:
            # Regular task - check if due_date matches
            if task.due_date == date:
                tasks_for_day.append({
                    'task': task,
                    'type': 'regular_task'
                })

    # Filter cultural tasks that occur on this date
    cultural_tasks_for_day = []
    for cultural_task in all_cultural_tasks:
        if cultural_task.is_recurring:
            # Check if this date is in the task's occurrence dates
            occurrences = cultural_task.get_occurrence_dates(date, date)
            if date in occurrences:
                cultural_tasks_for_day.append(cultural_task)
        else:
            # Non-recurring cultural task - check if due_date matches
            if cultural_task.due_date == date:
                cultural_tasks_for_day.append(cultural_task)

    # Filter sports tasks that occur on this date
    sports_tasks_for_day = []
    for sports_task in all_sports_tasks:
        if sports_task.is_recurring:
            # Check if this date is in the task's occurrence dates
            occurrences = sports_task.get_occurrence_dates(date, date)
            if date in occurrences:
                sports_tasks_for_day.append(sports_task)
        else:
            # Non-recurring sports task - check if due_date matches
            if sports_task.due_date == date:
                sports_tasks_for_day.append(sports_task)

    sharia_tasks_for_day = []
    for sharia_task in all_sharia_tasks:
        if sharia_task.is_recurring:
            # Check if this date is in the task's occurrence dates
            occurrences = sharia_task.get_occurrence_dates(date, date)
            if date in occurrences:
                sharia_tasks_for_day.append(sharia_task)
        else:
            # Non-recurring sharia task - check if due_date matches
            if sharia_task.due_date == date:
                sharia_tasks_for_day.append(sharia_task)

    scientific_tasks_for_day = []
    for scientific_task in all_scientific_tasks:
        if scientific_task.is_recurring:
            # Check if this date is in the task's occurrence dates
            occurrences = scientific_task.get_occurrence_dates(date, date)
            if date in occurrences:
                scientific_tasks_for_day.append(scientific_task)
        else:
            # Non-recurring scientific task - check if due_date matches
            if scientific_task.due_date == date:
                scientific_tasks_for_day.append(scientific_task)

    # Convert to queryset-like behavior for template
    tasks = [item['task'] for item in tasks_for_day if item['type'] == 'regular_task']
    cultural_tasks = cultural_tasks_for_day
    sharia_tasks = sharia_tasks_for_day
    sports_tasks = sports_tasks_for_day
    scientific_tasks = scientific_tasks_for_day

    activities = Activity.objects.filter(
        program=program,
        date=date
    ).select_related('committee', 'created_by')

    task_sessions = TaskSession.objects.filter(
        task__committee__program=program,
        date=date
    ).select_related('task', 'task__committee')

    operations_tasks = OperationsTask.objects.filter(
        committee__program=program,
        due_date=date
    ).select_related('committee')

    lectures = Lecture.objects.filter(
        committee__program=program,
        date=date
    ).select_related('committee', 'created_by')

    family_competitions = FamilyCompetition.objects.filter(
        committee__program=program,
        start_date__lte=date,
        end_date__gte=date
    ).select_related('committee')

    matches = Match.objects.filter(
        committee__program=program,
        date=date
    ).select_related('committee', 'created_by')

    # Filter by committee if supervisor
    if user.role == 'committee_supervisor':
        try:
            committee = Committee.objects.get(supervisor=user)
            events = events.filter(Q(committee=committee) | Q(committee__isnull=True))
            tasks = [t for t in tasks if t.committee == committee or t.committee is None]
            activities = activities.filter(Q(committee=committee) | Q(committee__isnull=True))
            cultural_tasks = [ct for ct in cultural_tasks if ct.committee == committee]
            sports_tasks = [st for st in sports_tasks if st.committee == committee]
            task_sessions = task_sessions.filter(task__committee=committee)
            operations_tasks = operations_tasks.filter(committee=committee)
            scientific_tasks = [st for st in scientific_tasks if st.committee == committee]
            lectures = lectures.filter(committee=committee)
            sharia_tasks = [st for st in sharia_tasks if st.committee == committee]
            family_competitions = family_competitions.filter(committee=committee)
            matches = matches.filter(committee=committee)
        except Committee.DoesNotExist:
            pass

    # Determine base template
    if user.role == 'director':
        base_template = 'director_base.html'
    elif user.role == 'program_manager':
        base_template = 'program_manager_base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'cultural':
        base_template = 'cultural_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'scientific':
        base_template = 'scientific_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'sports':
        base_template = 'sports_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'sharia':
        base_template = 'sharia_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'operations':
        base_template = 'operations_committee/base.html'
    else:
        base_template = 'base.html'

    context = {
        'program': program,
        'date': date,
        'events': events,
        'tasks': tasks,
        'tasks_count': len(tasks),
        'activities': activities,
        'cultural_tasks': cultural_tasks,
        'cultural_tasks_count': len(cultural_tasks),
        'sports_tasks': sports_tasks,
        'sports_tasks_count': len(sports_tasks),
        'task_sessions': task_sessions,
        'operations_tasks': operations_tasks,
        'scientific_tasks': scientific_tasks,
        'scientific_tasks_count': len(scientific_tasks),
        'lectures': lectures,
        'sharia_tasks': sharia_tasks,
        'sharia_tasks_count': len(sharia_tasks),
        'family_competitions': family_competitions,
        'matches': matches,
        'base_template': base_template,
    }

    return render(request, 'schedule/day_events.html', context)


def has_permission_for_program(user, program):
    """Check if user has permission for this program"""
    if user.role == 'director':
        return True
    elif user.role == 'program_manager':
        return program.manager == user
    elif user.role == 'committee_supervisor':
        try:
            committee = Committee.objects.get(supervisor=user)
            return committee.program == program
        except Committee.DoesNotExist:
            return False
    return False

@login_required
def object_list(request, object_type, program_id):
    """List view for different object types"""
    program = get_object_or_404(Program, id=program_id)

    # Check permissions
    if not has_permission_for_program(request.user, program):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('home')

    model_map = {
        'task': (Task, 'tasks'),
        'activity': (Activity, 'activities'),
        'cultural_task': (CulturalTask, 'cultural_tasks'),
        'cultural_report': (CulturalReport, 'cultural_reports'),
    }

    if object_type not in model_map:
        messages.error(request, 'نوع الكائن غير صحيح')
        return redirect('home')

    model_class, template_name = model_map[object_type]

    if object_type == 'cultural_task' or object_type == 'cultural_report':
        objects = model_class.objects.filter(committee__program=program)
    else:
        objects = model_class.objects.filter(program=program)

    context = {
        'objects': objects,
        'program': program,
        'object_type': object_type,
        'object_type_display': get_object_type_display(object_type),
    }

    return render(request, f'schedule/{template_name}_list.html', context)





@login_required
def event_detail(request, event_id):
    """Event detail page"""
    event = get_object_or_404(ScheduleEvent, id=event_id)
    user = request.user

    # Check permissions
    if user.role == 'director':
        pass  # Can view all
    elif user.role == 'program_manager':
        if event.program.manager != user:
            messages.error(request, 'ليس لديك صلاحية للوصول إلى هذا الحدث')
            return redirect('home')
    elif user.role == 'committee_supervisor':
        try:
            committee = Committee.objects.get(supervisor=user)
            if event.program != committee.program:
                messages.error(request, 'ليس لديك صلاحية للوصول إلى هذا الحدث')
                return redirect('home')
        except Committee.DoesNotExist:
            messages.error(request, 'لم يتم تعيين لجنة لك بعد')
            return redirect('home')

    attendances = event.attendances.select_related('user', 'recorded_by').all()

    context = {
        'event': event,
        'attendances': attendances,
    }

    return render(request, 'schedule/event_detail.html', context)


@login_required
def add_event(request, program_id=None):
    """Add new event"""
    user = request.user

    # Get program
    if user.role == 'director':
        if not program_id:
            messages.error(request, 'يجب تحديد برنامج')
            return redirect('home')
        program = get_object_or_404(Program, id=program_id)
    elif user.role == 'program_manager':
        try:
            program = Program.objects.get(manager=user)
        except Program.DoesNotExist:
            messages.error(request, 'لم يتم تعيين برنامج لك بعد')
            return redirect('home')
    elif user.role == 'committee_supervisor':
        try:
            committee = Committee.objects.get(supervisor=user)
            program = committee.program
        except Committee.DoesNotExist:
            messages.error(request, 'لم يتم تعيين لجنة لك بعد')
            return redirect('home')
    else:
        messages.error(request, 'ليس لديك صلاحية لإضافة أحداث')
        return redirect('home')

    if request.method == 'POST':
        form = ScheduleEventForm(request.POST, program=program)
        if form.is_valid():
            event = form.save(commit=False)
            event.program = program
            event.created_by = user
            event.save()


            UserActivity.objects.create(
                user=user,
                action=f'إضافة حدث جدولة: {event.title}',
                ip_address=get_client_ip(request)
            )

            messages.success(request, 'تم إضافة الحدث بنجاح!')
            return redirect('schedule_calendar', program_id=program.id)
    else:
        form = ScheduleEventForm(program=program)

    context = {
        'form': form,
        'program': program,
        'title': 'إضافة حدث جديد'
    }
    return render(request, 'schedule/event_form.html', context)



@login_required
def edit_event(request, event_id):
    """Edit existing event"""
    event = get_object_or_404(ScheduleEvent, id=event_id)
    user = request.user

    # Check permissions
    if user.role not in ['director', 'program_manager', 'committee_supervisor']:
        messages.error(request, 'ليس لديك صلاحية لتعديل الأحداث')
        return redirect('home')

    if request.method == 'POST':
        form = ScheduleEventForm(request.POST, instance=event, program=event.program)
        if form.is_valid():
            event = form.save()

            UserActivity.objects.create(
                user=user,
                action=f'تعديل حدث جدولة: {event.title}',
                ip_address=get_client_ip(request)
            )

            messages.success(request, 'تم تعديل الحدث بنجاح!')
            return redirect('event_detail', event_id=event.id)
    else:
        form = ScheduleEventForm(instance=event, program=event.program)

    context = {
        'form': form,
        'event': event,
        'program': event.program,
        'title': 'تعديل الحدث'
    }
    return render(request, 'schedule/event_form.html', context)


@login_required
def delete_event(request, event_id):
    """Delete event"""
    event = get_object_or_404(ScheduleEvent, id=event_id)
    user = request.user

    # Check permissions
    if user.role not in ['director', 'program_manager']:
        messages.error(request, 'ليس لديك صلاحية لحذف الأحداث')
        return redirect('home')

    program_id = event.program.id

    if request.method == 'POST':
        event_title = event.title
        event.delete()

        UserActivity.objects.create(
            user=user,
            action=f'حذف حدث جدولة: {event_title}',
            ip_address=get_client_ip(request)
        )

        messages.success(request, 'تم حذف الحدث بنجاح!')
        return redirect('schedule_calendar', program_id=program_id)

    context = {
        'object': event,
        'type': 'حدث',
        'program': event.program
    }
    return render(request, 'schedule/confirm_delete.html', context)


# Add these imports at the top of views.py
from pm_dashboard.models import Task, Activity,StudentAttendance
from cultural_committee_dashboard.models import CulturalTask, CulturalReport, FileLibrary, Discussion,CommitteeMember,DiscussionComment


@login_required
def object_detail(request, object_type, object_id):
    """Generic detail page for any object type"""
    user = request.user

    # Map object types to their models
    model_map = {
        'task': Task,
        'activity': Activity,
        'cultural_task': CulturalTask,
        'cultural_report': CulturalReport,
        'file': FileLibrary,
        'discussion': Discussion,
        'operations_task': OperationsTask,
        'scientific_task': ScientificTask,
        'lecture': Lecture,
        'sharia_task': ShariaTask,
        'family_competition': FamilyCompetition,
        'sports_task': SportsTask,
        'match': Match,
    }

    if object_type not in model_map:
        messages.error(request, 'نوع الكائن غير صحيح')
        return redirect('home')

    model_class = model_map[object_type]
    obj = get_object_or_404(model_class, id=object_id)

    # Check permissions based on object type and user role
    if not has_permission_for_object(user, obj):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('home')

    # Determine base template based on user role
    if user.role == 'director':
        base_template = 'director_base.html'
    elif user.role == 'program_manager':
        base_template = 'program_manager_base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'cultural':
        base_template = 'cultural_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'operations':
        base_template = 'operations_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'scientific':
        base_template = 'scientific_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'sharia':
        base_template = 'sharia_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'sports':
        base_template = 'sports_committee/base.html'
    else:
        base_template = 'base.html'

    # Get template context based on object type
    context = get_object_context(obj, object_type)
    context.update({
        'object': obj,
        'object_type': object_type,
        'object_type_display': get_object_type_display(object_type),
        'base_template': base_template,
    })

    template_name = get_template_name(object_type)
    return render(request, template_name, context)


def has_permission_for_object(user, obj):
    """Check if user has permission to view this object"""
    # Director can view everything
    if user.role == 'director':
        return True

    # Program Manager can view objects in their program
    if user.role == 'program_manager':
        program = get_object_program(obj)
        return program and program.manager == user

    # Committee Supervisor can view objects in their committee/program
    if user.role == 'committee_supervisor':
        program = get_object_program(obj)
        committee = get_object_committee(obj)

        if program:
            try:
                user_committee = Committee.objects.get(supervisor=user)
                return user_committee.program == program
            except Committee.DoesNotExist:
                return False

        if committee:
            return committee.supervisor == user

    return False


def get_object_program(obj):
    """Extract program from different object types"""
    if hasattr(obj, 'program'):
        return obj.program
    elif hasattr(obj, 'committee'):
        return obj.committee.program if obj.committee else None
    return None


def get_object_committee(obj):
    """Extract committee from different object types"""
    if hasattr(obj, 'committee'):
        return obj.committee
    return None


def get_object_context(obj, object_type):
    """Get specific context for each object type"""
    context = {}

    if object_type == 'task':
        context.update({
            'is_overdue': obj.is_overdue,
            'related_activities': Activity.objects.filter(program=obj.program, committee=obj.committee),
        })

    elif object_type == 'activity':
        context.update({
            'attendances': StudentAttendance.objects.filter(activity=obj).select_related('student', 'recorded_by'),
            'total_participants': obj.attendances.count(),
            'attended_count': obj.attendances.filter(attended=True).count(),
        })

    elif object_type == 'cultural_task':
        from cultural_committee_dashboard.models import CommitteeMember, FileLibrary
        context.update({
            'committee_members': CommitteeMember.objects.filter(committee=obj.committee, is_active=True),
            'related_files': FileLibrary.objects.filter(committee=obj.committee, file_type='cultural_plan'),
        })

    elif object_type == 'operations_task':
        from operations_committee_dashboard.models import OperationsTeamMember, LogisticsResource
        context.update({
            'committee_members': OperationsTeamMember.objects.filter(committee=obj.committee, is_active=True),
            'related_resources': LogisticsResource.objects.filter(committee=obj.committee),
            'is_overdue': obj.is_overdue,
        })

    elif object_type == 'scientific_task':
        from scientific_committee_dashboard.models import ScientificMember, ScientificFile
        context.update({
            'committee_members': ScientificMember.objects.filter(committee=obj.committee, is_active=True),
            'related_files': ScientificFile.objects.filter(committee=obj.committee),
        })

    elif object_type == 'lecture':
        from scientific_committee_dashboard.models import LectureAttendance
        context.update({
            'attendances': LectureAttendance.objects.filter(lecture=obj).select_related('user', 'recorded_by'),
            'total_participants': obj.attendances.count(),
            'attended_count': obj.attendances.filter(attended=True).count(),
        })

    elif object_type == 'sharia_task':
        from sharia_committee_dashboard.models import ShariaMember, ShariaFile
        context.update({
            'committee_members': ShariaMember.objects.filter(committee=obj.committee, is_active=True),
            'related_files': ShariaFile.objects.filter(committee=obj.committee),
        })

    elif object_type == 'family_competition':
        context.update({
            'is_active': obj.status == 'active',
            'is_upcoming': obj.status == 'upcoming',
        })

    elif object_type == 'sports_task':
        from sports_committee_dashboard.models import SportsMember
        context.update({
            'committee_members': SportsMember.objects.filter(committee=obj.committee, is_active=True),
        })

    elif object_type == 'match':
        context.update({
            'is_completed': obj.status == 'completed',
            'has_scores': obj.team1_score is not None and obj.team2_score is not None,
        })

    return context


def get_object_type_display(object_type):
    """Get display name for object type"""
    display_names = {
        'task': 'مهمة',
        'activity': 'نشاط',
        'cultural_task': 'مهمة ثقافية',
        'cultural_report': 'تقرير ثقافي',
        'file': 'ملف',
        'discussion': 'نقاش',
        'operations_task': 'مهمة تشغيلية',
        'scientific_task': 'مهمة علمية',
        'lecture': 'محاضرة',
        'sharia_task': 'مهمة شرعية',
        'family_competition': 'مسابقة أسرية',
        'sports_task': 'مهمة رياضية',
        'match': 'مباراة',
    }
    return display_names.get(object_type, 'كائن')


def get_template_name(object_type):
    """Get template name for each object type"""
    return f'schedule/{object_type}_detail.html'


from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator
from director_dashboard.models import DirectorAlbum, AlbumPhoto


def public_albums(request):
    """Public albums page for non-authenticated users"""
    # Get only active albums
    albums = DirectorAlbum.objects.filter(is_active=True).prefetch_related('photos').order_by('-created_at')

    # Pagination
    paginator = Paginator(albums, 9)  # Show 9 albums per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'albums': page_obj,
        'total_albums': albums.count(),
    }
    return render(request, 'main/albums.html', context)


def public_album_detail(request, album_id):
    """Public album detail page showing all photos"""
    album = get_object_or_404(DirectorAlbum, id=album_id, is_active=True)
    photos = album.photos.all().order_by('order', '-created_at')

    # Pagination for photos
    paginator = Paginator(photos, 12)  # Show 12 photos per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'album': album,
        'photos': page_obj,
        'total_photos': photos.count(),
    }
    return render(request, 'main/album_detail.html', context)


from django.http import HttpResponse
from datetime import datetime, timedelta
import csv
from io import BytesIO

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@login_required
def export_calendar_ics(request):
    """Export calendar as ICS file"""
    program_id = request.GET.get('program_id')
    view_type = request.GET.get('view', 'monthly')
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    week = int(request.GET.get('week', timezone.now().isocalendar()[1]))

    program = get_object_or_404(Program, id=program_id)

    # Check permissions
    if not has_permission_for_program(request.user, program):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذا التقويم')
        return redirect('home')

    # Get date range
    if view_type == 'weekly':
        jan_1 = datetime(year, 1, 1).date()
        start_date = jan_1 + timedelta(weeks=week - 1)
        start_date = start_date - timedelta(days=(start_date.weekday() + 1) % 7)
        end_date = start_date + timedelta(days=6)
    else:
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, month, monthrange(year, month)[1]).date()

    # Get all events
    events = ScheduleEvent.objects.filter(
        program=program,
        start_date__gte=start_date,
        start_date__lte=end_date
    ).select_related('committee', 'created_by')

    # Filter by committee if supervisor
    if request.user.role == 'committee_supervisor':
        try:
            committee = Committee.objects.get(supervisor=request.user)
            events = events.filter(Q(committee=committee) | Q(committee__isnull=True))
        except Committee.DoesNotExist:
            pass

    # Create ICS content
    response = HttpResponse(content_type='text/calendar; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="calendar_{program.name}_{year}_{month}.ics"'

    ics_content = ["BEGIN:VCALENDAR"]
    ics_content.append("VERSION:2.0")
    ics_content.append(f"PRODID:-//{program.name}//Calendar//AR")
    ics_content.append("CALSCALE:GREGORIAN")
    ics_content.append("METHOD:PUBLISH")

    for event in events:
        ics_content.append("BEGIN:VEVENT")

        # Format dates
        dtstart = event.start_date.strftime('%Y%m%d')
        if event.start_time:
            dtstart += f"T{event.start_time.strftime('%H%M%S')}"
        ics_content.append(f"DTSTART:{dtstart}")

        if event.end_date:
            dtend = event.end_date.strftime('%Y%m%d')
            if event.end_time:
                dtend += f"T{event.end_time.strftime('%H%M%S')}"
            ics_content.append(f"DTEND:{dtend}")

        ics_content.append(f"SUMMARY:{event.title}")
        ics_content.append(f"DESCRIPTION:{event.description}")

        if event.location:
            ics_content.append(f"LOCATION:{event.location}")

        ics_content.append(f"UID:{event.id}@{program.name}")
        ics_content.append(f"DTSTAMP:{timezone.now().strftime('%Y%m%dT%H%M%SZ')}")
        ics_content.append("END:VEVENT")

    ics_content.append("END:VCALENDAR")

    response.write('\r\n'.join(ics_content))
    return response


@login_required
def export_calendar_excel(request):
    """Export calendar as Excel file"""
    if not OPENPYXL_AVAILABLE:
        messages.error(request, 'مكتبة Excel غير متوفرة. الرجاء تثبيت openpyxl')
        return redirect('schedule_calendar')

    program_id = request.GET.get('program_id')
    view_type = request.GET.get('view', 'monthly')
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    week = int(request.GET.get('week', timezone.now().isocalendar()[1]))

    program = get_object_or_404(Program, id=program_id)

    # Check permissions
    if not has_permission_for_program(request.user, program):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذا التقويم')
        return redirect('home')

    # Get date range
    if view_type == 'weekly':
        jan_1 = datetime(year, 1, 1).date()
        start_date = jan_1 + timedelta(weeks=week - 1)
        start_date = start_date - timedelta(days=(start_date.weekday() + 1) % 7)
        end_date = start_date + timedelta(days=6)
    else:
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, month, monthrange(year, month)[1]).date()

    # Get all events
    events = ScheduleEvent.objects.filter(
        program=program,
        start_date__gte=start_date,
        start_date__lte=end_date
    ).select_related('committee', 'created_by').order_by('start_date', 'start_time')

    tasks = Task.objects.filter(
        program=program,
        due_date__gte=start_date,
        due_date__lte=end_date
    ).select_related('committee').order_by('due_date')

    activities = Activity.objects.filter(
        program=program,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('committee').order_by('date')

    # Filter by committee if supervisor
    if request.user.role == 'committee_supervisor':
        try:
            committee = Committee.objects.get(supervisor=request.user)
            events = events.filter(Q(committee=committee) | Q(committee__isnull=True))
            tasks = tasks.filter(Q(committee=committee) | Q(committee__isnull=True))
            activities = activities.filter(Q(committee=committee) | Q(committee__isnull=True))
        except Committee.DoesNotExist:
            pass

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "التقويم"
    ws.right_to_left = True

    # Styles
    header_fill = PatternFill(start_color="0084AB", end_color="0084AB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["التاريخ", "النوع", "العنوان", "الوصف", "الوقت", "المكان", "اللجنة", "الحالة"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Add events
    row = 2
    for event in events:
        ws.cell(row=row, column=1, value=event.start_date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row, column=2, value=event.get_event_type_display()).border = border
        ws.cell(row=row, column=3, value=event.title).border = border
        ws.cell(row=row, column=4, value=event.description).border = border
        ws.cell(row=row, column=5, value=event.start_time.strftime('%H:%M') if event.start_time else '').border = border
        ws.cell(row=row, column=6, value=event.location or '').border = border
        ws.cell(row=row, column=7, value=event.committee.name if event.committee else '').border = border
        ws.cell(row=row, column=8, value=event.get_status_display()).border = border
        row += 1

    # Add tasks
    for task in tasks:
        ws.cell(row=row, column=1, value=task.due_date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row, column=2, value='مهمة').border = border
        ws.cell(row=row, column=3, value=task.title).border = border
        ws.cell(row=row, column=4, value=task.description).border = border
        ws.cell(row=row, column=5, value='').border = border
        ws.cell(row=row, column=6, value='').border = border
        ws.cell(row=row, column=7, value=task.committee.name if task.committee else '').border = border
        ws.cell(row=row, column=8, value=task.get_status_display()).border = border
        row += 1

    # Add activities
    for activity in activities:
        ws.cell(row=row, column=1, value=activity.date.strftime('%Y-%m-%d')).border = border
        ws.cell(row=row, column=2, value='نشاط').border = border
        ws.cell(row=row, column=3, value=activity.name).border = border
        ws.cell(row=row, column=4, value=activity.description or '').border = border
        ws.cell(row=row, column=5, value=activity.time.strftime('%H:%M') if activity.time else '').border = border
        ws.cell(row=row, column=6, value=activity.location or '').border = border
        ws.cell(row=row, column=7, value=activity.committee.name if activity.committee else '').border = border
        ws.cell(row=row, column=8, value='').border = border
        row += 1

    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="calendar_{program.name}_{year}_{month}.xlsx"'

    wb.save(response)
    return response


@login_required
def export_calendar_pdf(request):
    """Export calendar as PDF file with Arabic support - includes all event types"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from io import BytesIO
        import os
        from django.conf import settings

        REPORTLAB_AVAILABLE = True
    except ImportError as e:
        print(f"❌ ReportLab import error: {e}")
        messages.error(request, 'مكتبة PDF غير متوفرة. الرجاء تثبيت reportlab')
        return redirect('schedule_calendar')

    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        ARABIC_TEXT_SUPPORT = True
    except ImportError:
        ARABIC_TEXT_SUPPORT = False
        print("⚠ Arabic text support libraries not found")

    def process_arabic_text(text):
        """Process Arabic text for proper RTL display"""
        if not text or not isinstance(text, str):
            return text or ''

        if not ARABIC_TEXT_SUPPORT:
            return text

        try:
            reshaper = arabic_reshaper.ArabicReshaper()
            reshaped_text = reshaper.reshape(text)
            bidi_text = get_display(reshaped_text)
            return bidi_text
        except Exception as e:
            print(f"⚠ Arabic text processing error: {e}")
            return text

    def hex_to_rgb(hex_color):
        """Convert hex color to RGB tuple (0-1 scale)"""
        hex_color = hex_color.lstrip('#')
        if len(hex_color) == 3:
            hex_color = ''.join([c * 2 for c in hex_color])

        r = int(hex_color[0:2], 16) / 255.0
        g = int(hex_color[2:4], 16) / 255.0
        b = int(hex_color[4:6], 16) / 255.0
        return (r, g, b)

    # Get parameters
    program_id = request.GET.get('program_id')
    view_type = request.GET.get('view', 'monthly')
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))
    week = int(request.GET.get('week', timezone.now().isocalendar()[1]))

    program = get_object_or_404(Program, id=program_id)

    if not has_permission_for_program(request.user, program):
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذا التقويم')
        return redirect('home')

    # Get date range
    if view_type == 'weekly':
        jan_1 = datetime(year, 1, 1).date()
        start_date = jan_1 + timedelta(weeks=week - 1)
        start_date = start_date - timedelta(days=(start_date.weekday() + 1) % 7)
        end_date = start_date + timedelta(days=6)
    else:
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, month, monthrange(year, month)[1]).date()

    # Get all types of events/tasks
    all_items = []

    # Schedule Events
    events = ScheduleEvent.objects.filter(
        program=program,
        start_date__gte=start_date,
        start_date__lte=end_date
    ).select_related('committee', 'created_by').order_by('start_date', 'start_time')

    for event in events:
        all_items.append({
            'date': event.start_date,
            'type': 'حدث',
            'title': event.title,
            'description': event.description,
            'committee': event.committee.name if event.committee else '',
            'item_type': 'event',
            'event_obj': event
        })

    # Tasks
    tasks = Task.objects.filter(
        program=program,
        due_date__gte=start_date,
        due_date__lte=end_date
    ).select_related('committee').order_by('due_date')

    for task in tasks:
        all_items.append({
            'date': task.due_date,
            'type': 'مهمة',
            'title': task.title,
            'description': task.description,
            'committee': task.committee.name if task.committee else '',
            'item_type': 'task',
            'event_obj': task
        })

    # Activities
    activities = Activity.objects.filter(
        program=program,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('committee').order_by('date')

    for activity in activities:
        all_items.append({
            'date': activity.date,
            'type': 'نشاط',
            'title': activity.name,
            'description': activity.description or '',
            'committee': activity.committee.name if activity.committee else '',
            'item_type': 'activity',
            'event_obj': activity
        })

    # Cultural Tasks
    cultural_tasks = CulturalTask.objects.filter(
        committee__program=program,
        due_date__gte=start_date,
        due_date__lte=end_date
    ).select_related('committee')

    for task in cultural_tasks:
        all_items.append({
            'date': task.due_date,
            'type': 'مهمة ثقافية',
            'title': task.title,
            'description': task.description or '',
            'committee': task.committee.name if task.committee else '',
            'item_type': 'cultural_task',
            'event_obj': task
        })

    # Task Sessions
    task_sessions = TaskSession.objects.filter(
        task__committee__program=program,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('task', 'task__committee')

    for session in task_sessions:
        all_items.append({
            'date': session.date,
            'type': 'جلسة مهمة',
            'title': session.title or f"جلسة {session.task.title}",
            'description': session.description or '',
            'committee': session.task.committee.name if session.task and session.task.committee else '',
            'item_type': 'task_session',
            'event_obj': session
        })

    # Operations Tasks
    operations_tasks = OperationsTask.objects.filter(
        committee__program=program,
        due_date__gte=start_date,
        due_date__lte=end_date
    ).select_related('committee')

    for task in operations_tasks:
        all_items.append({
            'date': task.due_date,
            'type': 'مهمة تشغيلية',
            'title': task.title,
            'description': task.description or '',
            'committee': task.committee.name if task.committee else '',
            'item_type': 'operations_task',
            'event_obj': task
        })

    # Scientific Tasks
    scientific_tasks = ScientificTask.objects.filter(
        committee__program=program,
        due_date__gte=start_date,
        due_date__lte=end_date
    ).select_related('committee')

    for task in scientific_tasks:
        all_items.append({
            'date': task.due_date,
            'type': 'مهمة علمية',
            'title': task.title,
            'description': task.description or '',
            'committee': task.committee.name if task.committee else '',
            'item_type': 'scientific_task',
            'event_obj': task
        })

    # Lectures
    lectures = Lecture.objects.filter(
        committee__program=program,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('committee', 'created_by')

    for lecture in lectures:
        all_items.append({
            'date': lecture.date,
            'type': 'محاضرة',
            'title': lecture.title,
            'description': lecture.description or '',
            'committee': lecture.committee.name if lecture.committee else '',
            'item_type': 'lecture',
            'event_obj': lecture
        })

    # Sharia Tasks
    sharia_tasks = ShariaTask.objects.filter(
        committee__program=program,
        due_date__gte=start_date,
        due_date__lte=end_date
    ).select_related('committee')

    for task in sharia_tasks:
        all_items.append({
            'date': task.due_date,
            'type': 'مهمة شرعية',
            'title': task.title,
            'description': task.description or '',
            'committee': task.committee.name if task.committee else '',
            'item_type': 'sharia_task',
            'event_obj': task
        })

    # Family Competitions
    family_competitions = FamilyCompetition.objects.filter(
        committee__program=program,
        start_date__lte=end_date,
        end_date__gte=start_date
    ).select_related('committee')

    for comp in family_competitions:
        all_items.append({
            'date': comp.start_date,
            'type': 'مسابقة أسرية',
            'title': comp.title,
            'description': comp.description or '',
            'committee': comp.committee.name if comp.committee else '',
            'item_type': 'family_competition',
            'event_obj': comp
        })

    # Sports Tasks
    sports_tasks = SportsTask.objects.filter(
        committee__program=program,
        due_date__gte=start_date,
        due_date__lte=end_date
    ).select_related('committee')

    for task in sports_tasks:
        all_items.append({
            'date': task.due_date,
            'type': 'مهمة رياضية',
            'title': task.title,
            'description': task.description or '',
            'committee': task.committee.name if task.committee else '',
            'item_type': 'sports_task',
            'event_obj': task
        })

    # Matches
    matches = Match.objects.filter(
        committee__program=program,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('committee', 'created_by')

    for match in matches:
        all_items.append({
            'date': match.date,
            'type': 'مباراة',
            'title': match.title,
            'description': match.description or '',
            'committee': match.committee.name if match.committee else '',
            'item_type': 'match',
            'event_obj': match
        })

    # Filter by committee if supervisor
    if request.user.role == 'committee_supervisor':
        try:
            committee = Committee.objects.get(supervisor=request.user)
            filtered_items = []
            for item in all_items:
                if item['committee'] == committee.name:
                    filtered_items.append(item)
            all_items = filtered_items
        except Committee.DoesNotExist:
            all_items = []

    # Sort all items by date
    all_items.sort(key=lambda x: x['date'])

    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="calendar_{program.name}_{year}_{month}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm
    )

    # Try to register Arabic font
    arabic_font = 'Helvetica'
    try:
        font_paths = [
            os.path.join(settings.BASE_DIR, 'main', 'static', 'fonts', 'IBMPlexSansArabic-Regular.ttf'),
            '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
            r'C:\Windows\Fonts\arial.ttf',
        ]

        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('IBMPlexArabic', font_path))
                    arabic_font = 'IBMPlexArabic'
                    break
                except Exception as e:
                    continue
    except Exception as e:
        arabic_font = 'Helvetica'

    # Create styles
    styles = getSampleStyleSheet()

    primary_color = (0, 0.33, 0.67)

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=primary_color,
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName=arabic_font,
    )

    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Normal'],
        fontSize=11,
        textColor=(0, 0, 0),
        alignment=TA_CENTER,
        fontName=arabic_font,
        leading=14,
    )

    cell_style = ParagraphStyle(
        'CustomCell',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        fontName=arabic_font,
        leading=12,
    )

    legend_style = ParagraphStyle(
        'LegendStyle',
        parent=styles['Normal'],
        fontSize=12,
        alignment=TA_RIGHT,
        fontName=arabic_font,
        spaceAfter=5,
    )

    # Build story
    story = []

    # Title
    title_text = process_arabic_text(f'تقويم {program.name} - {year}/{month}')
    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 0.3 * cm))


    # Legend items as simple text
    legend_items = [
        ('اجتماع', '#8B5CF6'),
        ('مبيت', '#EC4899'),
        ('رحلة', '#10B981'),
        ('الثقافية', '#3B82F6'),
        ('الرياضية', '#0EA5E9'),
        ('العلمية', '#FB923C'),
        ('الشرعية', '#F59E0B'),
        ('التشغيلية', '#8B5A2B'),
        ('المهمات', '#6B7280'),
        ('النشاطات', '#10B981'),
        ('المباريات', '#EF4444'),
        ('المحاضرات', '#8B5CF6'),
        ('المسابقات', '#EC4899'),
    ]

    # Build table data with only 5 columns: التاريخ, النوع, العنوان, الوصف, اللجنة
    table_data = []

    # Header row (only 5 columns now)
    headers = ['التاريخ', 'النوع', 'العنوان', 'الوصف', 'اللجنة']
    table_data.append([Paragraph(process_arabic_text(header), header_style) for header in headers])

    # Add all items to table
    for item in all_items:
        row = []

        # Date
        date_str = item['date'].strftime('%Y-%m-%d')
        row.append(Paragraph(process_arabic_text(date_str), cell_style))

        # Type
        row.append(Paragraph(process_arabic_text(item['type']), cell_style))

        # Title
        title = item['title'][:30] + '...' if len(item['title']) > 30 else item['title']
        row.append(Paragraph(process_arabic_text(title), cell_style))

        # Description
        desc = item['description'][:40] + '...' if len(item['description']) > 40 else item['description']
        row.append(Paragraph(process_arabic_text(desc), cell_style))

        # Committee
        row.append(Paragraph(process_arabic_text(item['committee']), cell_style))

        table_data.append(row)

    # Create table with 5 columns
    col_widths = [2.5 * cm, 2.5 * cm, 4.5 * cm, 5 * cm, 3 * cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Style table
    table_style = [
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), (0.21, 0.38, 0.57)),
        ('TEXTCOLOR', (0, 0), (-1, 0), (1, 1, 1)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), arabic_font),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),

        # Data rows
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, (0.7, 0.7, 0.7)),
        ('FONTNAME', (0, 1), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]

    # Add alternating row colors for better readability
    for row_idx in range(1, len(table_data)):
        if row_idx % 2 == 0:
            table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), (0.95, 0.95, 0.95)))

    table.setStyle(TableStyle(table_style))
    story.append(table)

    # Add summary
    story.append(Spacer(1, 0.5 * cm))
    summary_text = process_arabic_text(f'إجمالي العناصر: {len(all_items)}')
    summary_style = ParagraphStyle(
        'SummaryStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        fontName=arabic_font,
        spaceBefore=10,
    )
    story.append(Paragraph(summary_text, summary_style))

    # Build PDF
    doc.build(story)

    return response

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Q, Count
from datetime import datetime, timedelta
from pm_dashboard.models import Task, Activity
from cultural_committee_dashboard.models import CulturalTask
from scientific_committee_dashboard.models import ScientificTask
from operations_committee_dashboard.models import OperationsTask
from sharia_committee_dashboard.models import ShariaTask
from sports_committee_dashboard.models import SportsTask


@login_required
def calendar_list_view(request):
    """
    Enhanced list view for all calendar tasks, program tasks, and committee tasks
    Including recurring tasks and all occurrences like in schedule_calendar
    """
    user = request.user

    # Get program based on user role (same logic as schedule_calendar)
    if user.role == 'director':
        program_id = request.GET.get('program_id')
        if not program_id:
            programs = Program.objects.all()
            if programs.count() == 1:
                program = programs.first()
            else:
                # Redirect to program selection or show all
                return redirect('schedule_calendar')
        else:
            program = get_object_or_404(Program, id=program_id)

    elif user.role == 'program_manager':
        try:
            program = Program.objects.get(manager=user)
        except Program.DoesNotExist:
            messages.error(request, 'لم يتم تعيين برنامج لك بعد')
            return redirect('home')

    elif user.role == 'committee_supervisor':
        try:
            committee = Committee.objects.get(supervisor=user)
            program = committee.program
        except Committee.DoesNotExist:
            messages.error(request, 'لم يتم تعيين لجنة لك بعد')
            return redirect('home')
    else:
        messages.error(request, 'ليس لديك صلاحية للوصول إلى هذه الصفحة')
        return redirect('home')

    # Get date range from query params or default to current month
    today = timezone.now().date()

    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')

    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = today.replace(day=1)
    else:
        date_from = today.replace(day=1)

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            if today.month == 12:
                date_to = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                date_to = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    else:
        if today.month == 12:
            date_to = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            date_to = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

    # ==================== GET ALL TASKS WITH RECURRING SUPPORT ====================

    # Get regular tasks with recurring support
    tasks = Task.objects.filter(program=program).select_related('committee')

    # Process recurring tasks (same logic as schedule_calendar)
    task_occurrences = {}
    all_program_tasks = []  # For list view display

    for task in tasks:
        if task.is_recurring:
            task_start = task.start_date or task.due_date
            if task_start > date_to:
                continue

            if task.recurrence_end_date and task.recurrence_end_date < date_from:
                continue

            groups = task.get_consecutive_day_groups(date_from, date_to)
            for group_idx, (group_start, group_end) in enumerate(groups):
                group_id = f"task_{task.id}_group_{group_idx}"

                current = group_start
                while current <= group_end:
                    if date_from <= current <= date_to:
                        if current not in task_occurrences:
                            task_occurrences[current] = []

                        task_occurrences[current].append({
                            'task': task,
                            'is_start': current == group_start,
                            'is_end': current == group_end,
                            'group_id': group_id,
                            'group_start': group_start,
                            'group_end': group_end,
                            'span_days': (group_end - group_start).days + 1,
                            'type': 'regular_task',
                            'date': current  # Add date for list view
                        })

                        # Add to list view items
                        if current == group_start:  # Only add once per group
                            all_program_tasks.append({
                                'task': task,
                                'date': current,
                                'is_recurring': True,
                                'group_start': group_start,
                                'group_end': group_end
                            })
                    current += timedelta(days=1)
        else:
            # Non-recurring task
            if date_from <= task.due_date <= date_to:
                if task.due_date not in task_occurrences:
                    task_occurrences[task.due_date] = []
                task_occurrences[task.due_date].append({
                    'task': task,
                    'is_start': True,
                    'is_end': True,
                    'group_id': f"task_{task.id}_single",
                    'group_start': task.due_date,
                    'group_end': task.due_date,
                    'span_days': 1,
                    'type': 'regular_task',
                    'date': task.due_date
                })

                all_program_tasks.append({
                    'task': task,
                    'date': task.due_date,
                    'is_recurring': False
                })

    # Get cultural tasks with recurring support
    cultural_tasks = CulturalTask.objects.filter(
        committee__program=program
    ).select_related('committee')

    cultural_task_occurrences = {}
    all_cultural_tasks = []

    for cultural_task in cultural_tasks:
        if cultural_task.is_recurring:
            cultural_task_start = cultural_task.start_date or cultural_task.due_date
            if cultural_task_start > date_to:
                continue

            if cultural_task.recurrence_end_date and cultural_task.recurrence_end_date < date_from:
                continue

            groups = cultural_task.get_consecutive_day_groups(date_from, date_to)
            for group_idx, (group_start, group_end) in enumerate(groups):
                group_id = f"cultural_task_{cultural_task.id}_group_{group_idx}"

                current = group_start
                while current <= group_end:
                    if date_from <= current <= date_to:
                        if current not in cultural_task_occurrences:
                            cultural_task_occurrences[current] = []

                        cultural_task_occurrences[current].append({
                            'task': cultural_task,
                            'is_start': current == group_start,
                            'is_end': current == group_end,
                            'group_id': group_id,
                            'group_start': group_start,
                            'group_end': group_end,
                            'span_days': (group_end - group_start).days + 1,
                            'type': 'cultural_task',
                            'date': current
                        })

                        if current == group_start:
                            all_cultural_tasks.append({
                                'task': cultural_task,
                                'date': current,
                                'is_recurring': True,
                                'group_start': group_start,
                                'group_end': group_end
                            })
                    current += timedelta(days=1)
        else:
            # Non-recurring cultural task
            if date_from <= cultural_task.due_date <= date_to:
                if cultural_task.due_date not in cultural_task_occurrences:
                    cultural_task_occurrences[cultural_task.due_date] = []
                cultural_task_occurrences[cultural_task.due_date].append({
                    'task': cultural_task,
                    'is_start': True,
                    'is_end': True,
                    'group_id': f"cultural_task_{cultural_task.id}_single",
                    'group_start': cultural_task.due_date,
                    'group_end': cultural_task.due_date,
                    'span_days': 1,
                    'type': 'cultural_task',
                    'date': cultural_task.due_date
                })

                all_cultural_tasks.append({
                    'task': cultural_task,
                    'date': cultural_task.due_date,
                    'is_recurring': False
                })

    # Get sports tasks with recurring support
    sports_tasks = SportsTask.objects.filter(
        committee__program=program
    ).select_related('committee')

    sports_task_occurrences = {}
    all_sports_tasks = []

    for sports_task in sports_tasks:
        if sports_task.is_recurring:
            sports_task_start = sports_task.start_date or sports_task.due_date
            if sports_task_start > date_to:
                continue

            if sports_task.recurrence_end_date and sports_task.recurrence_end_date < date_from:
                continue

            groups = sports_task.get_consecutive_day_groups(date_from, date_to)
            for group_idx, (group_start, group_end) in enumerate(groups):
                group_id = f"sports_task_{sports_task.id}_group_{group_idx}"

                current = group_start
                while current <= group_end:
                    if date_from <= current <= date_to:
                        if current not in sports_task_occurrences:
                            sports_task_occurrences[current] = []

                        sports_task_occurrences[current].append({
                            'task': sports_task,
                            'is_start': current == group_start,
                            'is_end': current == group_end,
                            'group_id': group_id,
                            'group_start': group_start,
                            'group_end': group_end,
                            'span_days': (group_end - group_start).days + 1,
                            'type': 'sports_task',
                            'date': current
                        })

                        if current == group_start:
                            all_sports_tasks.append({
                                'task': sports_task,
                                'date': current,
                                'is_recurring': True,
                                'group_start': group_start,
                                'group_end': group_end
                            })
                    current += timedelta(days=1)
        else:
            # Non-recurring sports task
            if date_from <= sports_task.due_date <= date_to:
                if sports_task.due_date not in sports_task_occurrences:
                    sports_task_occurrences[sports_task.due_date] = []
                sports_task_occurrences[sports_task.due_date].append({
                    'task': sports_task,
                    'is_start': True,
                    'is_end': True,
                    'group_id': f"sports_task_{sports_task.id}_single",
                    'group_start': sports_task.due_date,
                    'group_end': sports_task.due_date,
                    'span_days': 1,
                    'type': 'sports_task',
                    'date': sports_task.due_date
                })

                all_sports_tasks.append({
                    'task': sports_task,
                    'date': sports_task.due_date,
                    'is_recurring': False
                })

    # Combine all task occurrences
    all_task_occurrences = {}
    for date, occurrences in task_occurrences.items():
        if date not in all_task_occurrences:
            all_task_occurrences[date] = []
        all_task_occurrences[date].extend(occurrences)

    for date, occurrences in cultural_task_occurrences.items():
        if date not in all_task_occurrences:
            all_task_occurrences[date] = []
        all_task_occurrences[date].extend(occurrences)

    for date, occurrences in sports_task_occurrences.items():
        if date not in all_task_occurrences:
            all_task_occurrences[date] = []
        all_task_occurrences[date].extend(occurrences)

    # Get activities
    activities = Activity.objects.filter(
        program=program,
        date__gte=date_from,
        date__lte=date_to
    ).select_related('committee', 'created_by').order_by('-date')

    # Get non-recurring cultural tasks separately for compatibility
    cultural_tasks_non_recurring = CulturalTask.objects.filter(
        committee__program=program,
        is_recurring=False,
        due_date__gte=date_from,
        due_date__lte=date_to
    ).select_related('committee').order_by('-due_date')

    # Get task sessions
    task_sessions = TaskSession.objects.filter(
        task__committee__program=program,
        date__gte=date_from,
        date__lte=date_to
    ).select_related('task', 'task__committee').order_by('-date')

    # Get operations tasks
    operations_tasks = OperationsTask.objects.filter(
        committee__program=program,
        due_date__gte=date_from,
        due_date__lte=date_to
    ).select_related('committee').order_by('-due_date')

    # Get scientific tasks
    scientific_tasks = ScientificTask.objects.filter(
        committee__program=program,
        due_date__gte=date_from,
        due_date__lte=date_to
    ).select_related('committee').order_by('-due_date')

    # Get lectures
    lectures = Lecture.objects.filter(
        committee__program=program,
        date__gte=date_from,
        date__lte=date_to
    ).select_related('committee', 'created_by').order_by('-date')

    # Get sharia tasks
    sharia_tasks = ShariaTask.objects.filter(
        committee__program=program,
        due_date__gte=date_from,
        due_date__lte=date_to
    ).select_related('committee').order_by('-due_date')

    # Get family competitions
    family_competitions = FamilyCompetition.objects.filter(
        committee__program=program,
        start_date__lte=date_to,
        end_date__gte=date_from
    ).select_related('committee').order_by('-start_date')

    # Get matches
    matches = Match.objects.filter(
        committee__program=program,
        date__gte=date_from,
        date__lte=date_to
    ).select_related('committee', 'created_by').order_by('-date')

    # Apply permission filters based on user role
    if user.role == 'committee_supervisor':
        try:
            committee = Committee.objects.get(supervisor=user)

            # Filter all by committee
            activities = activities.filter(committee=committee)
            cultural_tasks_non_recurring = cultural_tasks_non_recurring.filter(committee=committee)
            task_sessions = task_sessions.filter(task__committee=committee)
            operations_tasks = operations_tasks.filter(committee=committee)
            scientific_tasks = scientific_tasks.filter(committee=committee)
            lectures = lectures.filter(committee=committee)
            sharia_tasks = sharia_tasks.filter(committee=committee)
            family_competitions = family_competitions.filter(committee=committee)
            matches = matches.filter(committee=committee)

            # Filter program tasks by committee
            all_program_tasks = [item for item in all_program_tasks if item['task'].committee == committee]

            # Filter based on supervisor type
            if user.supervisor_type == 'cultural':
                all_cultural_tasks = [item for item in all_cultural_tasks if item['task'].committee == committee]
                all_sports_tasks = []
                operations_tasks = OperationsTask.objects.none()
                scientific_tasks = ScientificTask.objects.none()
                lectures = Lecture.objects.none()
                sharia_tasks = ShariaTask.objects.none()
                family_competitions = FamilyCompetition.objects.none()
                matches = Match.objects.none()
            elif user.supervisor_type == 'scientific':
                scientific_tasks = scientific_tasks.filter(committee=committee)
                lectures = lectures.filter(committee=committee)
                all_cultural_tasks = []
                all_sports_tasks = []
                operations_tasks = OperationsTask.objects.none()
                sharia_tasks = ShariaTask.objects.none()
                family_competitions = FamilyCompetition.objects.none()
                matches = Match.objects.none()
            elif user.supervisor_type == 'operations':
                operations_tasks = operations_tasks.filter(committee=committee)
                all_cultural_tasks = []
                all_sports_tasks = []
                scientific_tasks = ScientificTask.objects.none()
                lectures = Lecture.objects.none()
                sharia_tasks = ShariaTask.objects.none()
                family_competitions = FamilyCompetition.objects.none()
                matches = Match.objects.none()
            elif user.supervisor_type == 'sharia':
                sharia_tasks = sharia_tasks.filter(committee=committee)
                family_competitions = family_competitions.filter(committee=committee)
                all_cultural_tasks = []
                all_sports_tasks = []
                operations_tasks = OperationsTask.objects.none()
                scientific_tasks = ScientificTask.objects.none()
                lectures = Lecture.objects.none()
                matches = Match.objects.none()
            elif user.supervisor_type == 'sports':
                all_sports_tasks = [item for item in all_sports_tasks if item['task'].committee == committee]
                matches = matches.filter(committee=committee)
                all_cultural_tasks = []
                operations_tasks = OperationsTask.objects.none()
                scientific_tasks = ScientificTask.objects.none()
                lectures = Lecture.objects.none()
                sharia_tasks = ShariaTask.objects.none()
                family_competitions = FamilyCompetition.objects.none()
        except Exception:
            pass

    # Calculate statistics
    total_tasks = (
            len(all_program_tasks) +
            len(all_cultural_tasks) +
            len(all_sports_tasks) +
            operations_tasks.count() +
            scientific_tasks.count() +
            sharia_tasks.count()
    )

    # Completed tasks - need to check each task in the lists
    completed_count = 0
    for item in all_program_tasks:
        if item['task'].status == 'completed':
            completed_count += 1
    for item in all_cultural_tasks:
        if item['task'].status == 'completed':
            completed_count += 1
    for item in all_sports_tasks:
        if item['task'].status == 'completed':
            completed_count += 1

    completed_tasks = (
            completed_count +
            operations_tasks.filter(status='completed').count() +
            scientific_tasks.filter(status='completed').count() +
            sharia_tasks.filter(status='completed').count()
    )

    # Pending/In Progress tasks
    pending_count = 0
    for item in all_program_tasks:
        if item['task'].status in ['pending', 'in_progress']:
            pending_count += 1
    for item in all_cultural_tasks:
        if item['task'].status in ['pending', 'in_progress']:
            pending_count += 1
    for item in all_sports_tasks:
        if item['task'].status in ['pending', 'in_progress']:
            pending_count += 1

    pending_tasks = (
            pending_count +
            operations_tasks.filter(Q(status='pending') | Q(status='in_progress')).count() +
            scientific_tasks.filter(Q(status='pending') | Q(status='in_progress')).count() +
            sharia_tasks.filter(Q(status='pending') | Q(status='in_progress')).count()
    )

    # Overdue tasks
    overdue_count = 0
    for item in all_program_tasks:
        task = item['task']
        task_date = item.get('group_end', task.due_date) if item['is_recurring'] else task.due_date
        if task_date < today and task.status in ['pending', 'in_progress']:
            overdue_count += 1

    for item in all_cultural_tasks:
        task = item['task']
        task_date = item.get('group_end', task.due_date) if item['is_recurring'] else task.due_date
        if task_date < today and task.status in ['pending', 'in_progress']:
            overdue_count += 1

    for item in all_sports_tasks:
        task = item['task']
        task_date = item.get('group_end', task.due_date) if item['is_recurring'] else task.due_date
        if task_date < today and task.status in ['pending', 'in_progress']:
            overdue_count += 1

    overdue_tasks = (
            overdue_count +
            operations_tasks.filter(due_date__lt=today, status__in=['pending', 'in_progress']).count() +
            scientific_tasks.filter(due_date__lt=today, status__in=['pending', 'in_progress']).count() +
            sharia_tasks.filter(due_date__lt=today, status__in=['pending', 'in_progress']).count()
    )

    # Determine base template
    if user.role == 'director':
        base_template = 'director_base.html'
    elif user.role == 'program_manager':
        base_template = 'program_manager_base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'cultural':
        base_template = 'cultural_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'sports':
        base_template = 'sports_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'sharia':
        base_template = 'sharia_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'operations':
        base_template = 'operations_committee/base.html'
    elif user.role == 'committee_supervisor' and user.supervisor_type == 'scientific':
        base_template = 'scientific_committee/base.html'
    else:
        base_template = 'base.html'

    context = {
        # Task occurrences (for compatibility with calendar view)
        'all_task_occurrences': all_task_occurrences,

        # List view items (tasks with their dates)
        'program_tasks': all_program_tasks,
        'cultural_tasks': all_cultural_tasks,
        'sports_tasks': all_sports_tasks,

        # Other items
        'activities': activities,
        'cultural_tasks_non_recurring': cultural_tasks_non_recurring,
        'task_sessions': task_sessions,
        'operations_tasks': operations_tasks,
        'scientific_tasks': scientific_tasks,
        'lectures': lectures,
        'sharia_tasks': sharia_tasks,
        'family_competitions': family_competitions,
        'matches': matches,

        # Metadata
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
        'total_tasks': total_tasks,
        'completed_tasks': completed_tasks,
        'pending_tasks': pending_tasks,
        'overdue_tasks': overdue_tasks,
        'base_template': base_template,
        'program': program,
    }

    return render(request, 'schedule/calendar_list_view.html', context)